"""
Tests for the Pyclient.
"""
import logging
import os
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from dotenv import load_dotenv

from src.molgenis_emx2_pyclient import Client
from src.molgenis_emx2_pyclient.exceptions import SigninError, SignoutError, NoSuchSchemaException, \
    ReferenceException, PermissionDeniedException, PyclientException
from src.molgenis_emx2_pyclient.metadata import Schema

load_dotenv()
server_url = os.environ.get("MG_SERVER")
username = os.environ.get("MG_USERNAME")
password = os.environ.get("MG_PASSWORD")

RESOURCES_DIR = Path(__file__).parent / "resources"

def test_signin():
    """Tests the `signin` method."""
    with pytest.raises(SigninError) as excinfo:
        with Client(url=server_url) as client:
            client.signin(username+username, password)
    assert excinfo.value.msg.endswith("Sign in as 'adminadmin' failed: user or password unknown")

    with Client(url=server_url) as client:
        client.signin(username, password)

        assert client.signin_status == "success"


def test_signout():
    """Tests the `signout` method."""
    with Client(url=server_url) as client:
        with pytest.raises(SignoutError) as excinfo:
            client.signout()
        assert excinfo.value.msg == "Could not sign out as user is not signed in."

        client.signin(username, password)
        client.signout()
        assert client.signin_status == "signed out"


def test_status():
    """Tests the `status` property."""
    with Client(url=server_url) as client:
        status: str = client.status
    assert status.split("Host: ")[-1].startswith(server_url)
    assert status.split("User: ")[-1].startswith("anonymous")
    assert status.split("Status: ")[-1].startswith("signed out")
    assert "pet store" in status.split("Schemas: ")[-1]


def test_get_schemas():
    """Tests the `get_schemas` method."""
    with Client(url=server_url) as client:
        schemas = client.get_schemas()
    assert "pet store" in map(lambda schema: schema.id, schemas)


def test_set_token():
    """Tests the `set_token` method."""
    token = "SAMPLE TOKEN"
    with Client(url=server_url) as client:
        client.set_token(token)

        assert client._token == token



def test_upload_csv():
    """Tests the `upload_csv` method."""
    with Client(url=server_url) as client:
        client.signin(username, password)

        with pytest.raises(FileNotFoundError) as excinfo:
            client._upload_csv(file_path=Path("Pet.csv"), schema="pet store")
        assert str(excinfo.value) == "[Errno 2] No such file or directory: 'Pet.csv'"

        client._upload_csv(file_path=RESOURCES_DIR / "insert" / "Tag.csv", schema="pet store")
        client._upload_csv(file_path=RESOURCES_DIR / "insert" / "Pet.csv", schema="pet store")
        client._upload_csv(file_path=RESOURCES_DIR / "delete" / "Pet.csv", schema="pet store")
        client._upload_csv(file_path=RESOURCES_DIR / "delete" / "Tag.csv", schema="pet store")


@pytest.mark.asyncio
async def test_upload_file():
    """Tests the `upload_file` method."""
    with Client(url=server_url) as client:
        client.signin(username, password)

        # Upload without specifying schema
        with pytest.raises(NoSuchSchemaException) as excinfo:
            await client.upload_file(file_path=RESOURCES_DIR / "insert" / "Pet.csv")
        assert excinfo.value.msg == f"Specify the schema where the file should be uploaded."

        # Upload ZIP file
        pet_before = len(client.get_graphql(schema="pet store", table="Pet", columns=["name"]))
        tag_before = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        await client.upload_file(file_path=RESOURCES_DIR / "insert" / "pet store.zip", schema="pet store")
        pet_between = len(client.get_graphql(schema="pet store", table="Pet", columns=["name"]))
        tag_between = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))

        assert pet_between == pet_before + 2
        assert tag_between == tag_before + 2

        await client.upload_file(file_path=RESOURCES_DIR / "delete" / "pet store.zip", schema="pet store")
        pet_after = len(client.get_graphql(schema="pet store", table="Pet", columns=["name"]))
        tag_after = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))

        assert pet_after == pet_before
        assert tag_after == tag_before

        # Upload XLSX file
        pet_before = len(client.get_graphql(schema="pet store", table="Pet", columns=["name"]))
        tag_before = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        await client.upload_file(file_path=RESOURCES_DIR / "insert" / "pet store.xlsx", schema="pet store")
        pet_between = len(client.get_graphql(schema="pet store", table="Pet", columns=["name"]))
        tag_between = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))

        assert pet_between == pet_before + 2
        assert tag_between == tag_before + 2

        await client.upload_file(file_path=RESOURCES_DIR / "delete" / "pet store.xlsx", schema="pet store")
        pet_after = len(client.get_graphql(schema="pet store", table="Pet", columns=["name"]))
        tag_after = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))

        assert pet_after == pet_before
        assert tag_after == tag_before

        # Upload with unsupported file type
        with pytest.raises(NotImplementedError) as excinfo:
            await client.upload_file(file_path=RESOURCES_DIR / "insert" / "Pet.txt", schema="pet store")
        assert str(excinfo.value) == "Uploading files with extension '.txt' is not supported."



def test_truncate():
    """Tests the `truncate` method."""
    with Client(url=server_url) as client:
        client.signin(username, password)

        # Test truncate with ReferenceException
        with pytest.raises(ReferenceException) as excinfo:
            client.truncate(schema='pet store', table='Pet')
        assert excinfo.value.msg.startswith("Transaction failed: delete on table \"Pet\" violates foreign key constraint.")

        # Test correct running
        client.truncate(schema='pet store', table='User')
        users_after = len(client.get_graphql(schema="pet store", table="User", columns=["username"]))
        assert users_after == 0

        client.save_schema(table="User", name="pet store", file=RESOURCES_DIR / "petstore" / "User.csv")


def test_delete_records():
    """Tests the `delete_records` method."""

    # Test fail without editor rights
    with Client(url=server_url) as client:
        with pytest.raises(PermissionDeniedException) as excinfo:
            client.delete_records(table="Pet", schema="pet store", file=RESOURCES_DIR / "insert" / "Pet.csv")
        assert str(excinfo.value) == "Message: Transaction failed: permission denied.\n"

        client.signin(username, password)

        # Test fail without schema
        with pytest.raises(NoSuchSchemaException) as excinfo:
            client.delete_records(table="Pet", file=RESOURCES_DIR / "insert" / "Pet.csv")

        assert excinfo.value.msg == "Schema None not available."

        # Test fail without specifying file or data
        with pytest.raises(FileNotFoundError) as excinfo:
            client.delete_records(schema="pet store" , table="Pet")

        assert str(excinfo.value) == "No data to import. Specify a file location or a dataset."

        # Test delete with file
        tag_before = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        client.save_schema(name="pet store", table="Tag", file=RESOURCES_DIR / "insert" / "Tag.csv")
        tag_between = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))

        assert tag_between == tag_before + 2
        client.delete_records(schema="pet store" , table="Tag", file=RESOURCES_DIR / "insert" / "Tag.csv")

        tag_after = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        assert tag_after == tag_before

        # Test delete with data as list
        tags_df = pd.read_csv(RESOURCES_DIR / "insert" / "Tag.csv")
        tags_list = list(tags_df.to_dict(orient='index').values())

        tag_before = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        client.save_schema(name="pet store", table="Tag", file=RESOURCES_DIR / "insert" / "Tag.csv")
        tag_between = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))

        assert tag_between == tag_before + 2
        client.delete_records(schema="pet store" , table="Tag", data=tags_list)

        tag_after = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        assert tag_after == tag_before

        # Test delete with data as DataFrame
        tag_before = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        client.save_schema(name="pet store", table="Tag", file=RESOURCES_DIR / "insert" / "Tag.csv")
        tag_between = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))

        assert tag_between == tag_before + 2
        client.delete_records(schema="pet store" , table="Tag", data=tags_df)

        tag_after = len(client.get_graphql(schema="pet store", table="Tag", columns=["name"]))
        assert tag_after == tag_before

@pytest.mark.asyncio
async def test_export():
    """Tests the `export` method."""
    with Client(url=server_url) as client:
        client.signin(username, password)

        # Upload without specifying schema
        with pytest.raises(NoSuchSchemaException) as excinfo:
            await client.export(table="Pet", filename="pet.csv")
        assert excinfo.value.msg == "Schema None not available."

        # Test incorrect file name
        with pytest.raises(ValueError) as excinfo:
            await client.export(schema="pet store", table="Pet", filename="pet.txt")
        assert str(excinfo.value) == "File name must end with ('csv', 'xlsx', 'zip')"

        # Test CSV
        csv_data: BytesIO = await client.export(schema="pet store", table="Pet")
        df = pd.read_csv(csv_data)
        assert len(df.columns) == 8
        await client.export(schema="pet store", table="Pet", filename="pet.csv")
        assert (Path(__file__).parent.parent / "pet.csv").exists()
        (Path(__file__).parent.parent / "pet.csv").unlink()

        # Test ZIP
        zip_data: BytesIO = await client.export(schema="pet store")
        with zipfile.ZipFile(zip_data, 'r') as zf:
            file_names = zf.namelist()
        assert len(file_names) == 8
        await client.export(schema="pet store", filename="pet store.zip")
        assert (Path(__file__).parent.parent / "pet store.zip").exists()
        (Path(__file__).parent.parent / "pet store.zip").unlink()

        # Test XLSX table
        xlsx_data: BytesIO = await client.export(schema="pet store", table="Pet", as_excel=True)
        book = openpyxl.load_workbook(xlsx_data, data_only=True)
        assert len(book.sheetnames) == 1
        await client.export(schema="pet store", table="Pet", as_excel=True, filename="pet.xlsx")
        assert (Path(__file__).parent.parent / "pet.xlsx").exists()
        (Path(__file__).parent.parent / "pet.xlsx").unlink()

        # Test XLSX tables as sheets
        xlsx_data: BytesIO = await client.export(schema="pet store", as_excel=True)
        book = openpyxl.load_workbook(xlsx_data, data_only=True)
        assert len(book.sheetnames) == 8
        await client.export(schema="pet store", as_excel=True, filename="pet store.xlsx")
        assert (Path(__file__).parent.parent / "pet store.xlsx").exists()
        (Path(__file__).parent.parent / "pet store.xlsx").unlink()


def test_get_schema_metadata():
    """Tests the `get_schema_metadata` method."""

    with Client(url=server_url) as client:
        with pytest.raises(NoSuchSchemaException) as excinfo:
            client.get_schema_metadata("pet store 2")
        assert excinfo.value.msg == "Schema 'pet store 2' not available."

        pet_meta = client.get_schema_metadata("pet store")
        assert len(pet_meta.get("members")) == 5
        assert len(pet_meta.get("roles")) == 8
        assert len(pet_meta.get("settings")) == 1
        assert len(pet_meta.tables) == 5
        assert len(pet_meta.get_table('name', 'Pet').columns) == 15

@pytest.mark.asyncio
async def test_create_schema():
    """Tests the `create_schema` method."""

    with Client(url=server_url) as client:
        client.signin(username, password)

        # Test fail on existing name
        with pytest.raises(PyclientException) as excinfo:
            await client.create_schema(name="pet store")
        assert excinfo.value.msg == "Schema with name 'pet store' already exists."

        # Test description
        await client.create_schema(name="pet store 2", description="The second pet store.")
        schemas: list[Schema] = client.get_schemas()
        pet_meta: list[Schema] = [s for s in schemas if s.get('name') == "pet store 2"]
        assert len(pet_meta) == 1
        assert pet_meta[0].get('description') == "The second pet store."
        await client.delete_schema("pet store 2")

        # Test template
        await client.create_schema(name="pet store 2",
                                   description="The second pet store.",
                                   template="PET_STORE")
        schemas: list[Schema] = client.get_schemas()
        pet_meta: list[Schema] = [s for s in schemas if s.get('name') == "pet store 2"]
        assert len(pet_meta) == 1
        assert len(client.get_schema_metadata("pet store 2").tables) == 5
        await client.delete_schema("pet store 2")


        # Test include demo data
        await client.create_schema(name="pet store 2",
                                   description="The second pet store.",
                                   template="PET_STORE",
                                   include_demo_data=True)
        schemas: list[Schema] = client.get_schemas()
        pet_meta: list[Schema] = [s for s in schemas if s.get('name') == "pet store 2"]
        assert len(pet_meta) == 1
        assert len(client.get_schema_metadata("pet store 2").tables) == 5
        assert len(client.get(table="Pet", schema="pet store 2")) == 8
        await client.delete_schema("pet store 2")


@pytest.mark.asyncio
async def test_delete_schema():
    """Tests the `delete_schema` method."""

    with Client(url=server_url) as client:
        client.signin(username, password)

        with pytest.raises(NoSuchSchemaException) as excinfo:
            await client.delete_schema("pet store 2")
        assert excinfo.value.msg == "Schema 'pet store 2' not available."

        await client.create_schema("pet store 2")

        schemas_before = client.get_schemas()
        await client.delete_schema("pet store 2")
        assert len(client.get_schemas()) == len(schemas_before) - 1


@pytest.mark.asyncio
async def test_update_schema():
    """Tests the `update_schema` method."""

    with Client(url=server_url) as client:
        client.signin(username, password)

        # Fail on missing schema
        with pytest.raises(NoSuchSchemaException) as excinfo:
            client.update_schema("pet store 2")
        assert excinfo.value.msg == "Schema 'pet store 2' not available."

        await client.create_schema("pet store 2")
        client.update_schema("pet store 2", "The second pet store.")
        schemas = client.get_schemas()
        pet2_meta: list[Schema] = [s for s in schemas if s.get('name') == "pet store 2"]
        assert pet2_meta[0].get("description") == "The second pet store."

        await client.delete_schema("pet store 2")

@pytest.mark.asyncio
async def test_recreate_schema():
    """Tests the `recreate_schema` method."""

    with Client(url=server_url) as client:
        client.signin(username, password)

        # Fail on missing schema
        with pytest.raises(NoSuchSchemaException) as excinfo:
            await client.recreate_schema("pet store 2")
        assert excinfo.value.msg == "Schema 'pet store 2' not available."

        await client.create_schema("pet store 2")

        await client.recreate_schema(name="pet store 2", description="The second pet store.")
        schemas = client.get_schemas()
        pet2_meta: list[Schema] = [s for s in schemas if s.get('name') == "pet store 2"]
        assert pet2_meta[0].get("description") == "The second pet store."

        await client.recreate_schema(name="pet store 2",
                                     description="The second pet store.",
                                     template="PET_STORE")
        assert len(client.get_schema_metadata("pet store 2").tables) == 5

        await client.recreate_schema(name="pet store 2",
                                     description="The second pet store.",
                                     template="PET_STORE",
                                     include_demo_data=True)
        assert len(client.get(table="Pet", schema="pet store 2")) == 8

        await client.delete_schema("pet store 2")


def test_set_schema():
    """Tests the `set_schema` method."""

    with Client(url=server_url) as client:
        current_schema = client.default_schema
        assert current_schema is None

        with pytest.raises(NoSuchSchemaException) as excinfo:
            client.set_schema("pet store 2")
        assert excinfo.value.msg == "Schema 'pet store 2' not available."

        client.set_schema("pet store")
        assert client.default_schema == "pet store"


@pytest.mark.asyncio
async def test_report_task_progress(caplog):
    """Tests the `_report_task_progress` method."""

    with Client(url=server_url) as client:
        client.signin(username, password)

        file_path = RESOURCES_DIR / "delete" / "pet store.zip"
        api_url = f"{client.url}/pet store/api/zip?async=true"

        with open(file_path, 'rb') as file:
            response = client.session.post(
                url=api_url,
                files={'file': file}
            )
        process_id = response.json().get('id')

        caplog.set_level(logging.INFO)
        await client._report_task_progress(process_id)

        message_starts = [
            "Import from store",
            "    Modified 2 rows in Pet in ",
            "    Modified 2 rows in Tag in ",
            "Committing",
            "Completed task: Import csv file in "
        ]
        for cm in caplog.messages:
            assert any(map(lambda ms: cm.startswith(ms), message_starts))


def test_validate_graphql_response():
    """Tests the `_validate_graphql_response` method."""
    ...
